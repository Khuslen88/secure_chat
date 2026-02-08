import json
import os
import uuid
from datetime import datetime, timezone

from werkzeug.utils import secure_filename

from config import Config
from security import SecurityUtils


class KnowledgeBase:
    """Manages company knowledge base documents for RAG-style context injection.

    Documents are stored in data/knowledge_base/ with extracted text cached
    in data/knowledge_base/extracted/ as .txt files. A metadata index
    (index.json) tracks all documents.
    """

    INDEX_FILE = os.path.join(Config.KNOWLEDGE_BASE_DIR, "index.json")
    ALLOWED_EXTENSIONS = {".pdf", ".txt", ".docx", ".csv", ".xlsx"}

    def __init__(self):
        os.makedirs(Config.KNOWLEDGE_BASE_DIR, exist_ok=True)
        os.makedirs(Config.EXTRACTED_TEXT_DIR, exist_ok=True)
        if not os.path.exists(self.INDEX_FILE):
            self._write_index([])

    def _read_index(self):
        with open(self.INDEX_FILE, "r") as f:
            return json.load(f)

    def _write_index(self, index):
        with open(self.INDEX_FILE, "w") as f:
            json.dump(index, f, indent=2)

    def add_document(self, file_storage):
        """Ingest a document into the knowledge base.

        Returns (success, doc_id_or_error) tuple.
        """
        raw_name = secure_filename(file_storage.filename or "")
        if not raw_name:
            return False, "Invalid filename."

        ext = os.path.splitext(raw_name)[1].lower()
        if ext not in self.ALLOWED_EXTENSIONS:
            allowed = ", ".join(sorted(self.ALLOWED_EXTENSIONS))
            return False, f"Knowledge base only accepts: {allowed}"

        # Measure size
        file_storage.stream.seek(0, os.SEEK_END)
        file_size = file_storage.stream.tell()
        file_storage.stream.seek(0)

        valid, err = SecurityUtils.validate_file(raw_name, file_size, file_storage.stream)
        if not valid:
            return False, err

        # Save document
        doc_id = str(uuid.uuid4())
        save_name = f"{doc_id}{ext}"
        doc_path = os.path.join(Config.KNOWLEDGE_BASE_DIR, save_name)
        file_storage.save(doc_path)

        # Extract and cache text
        try:
            text = self.extract_text(doc_path, ext)
        except Exception as e:
            os.remove(doc_path)
            return False, f"Failed to extract text: {e}"

        text_path = os.path.join(Config.EXTRACTED_TEXT_DIR, f"{doc_id}.txt")
        with open(text_path, "w") as f:
            f.write(text)

        # Add to index
        entry = {
            "id": doc_id,
            "original_name": raw_name,
            "stored_name": save_name,
            "extension": ext,
            "size": file_size,
            "text_length": len(text),
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
        }
        index = self._read_index()
        index.append(entry)
        self._write_index(index)

        return True, doc_id

    def extract_text(self, filepath, ext):
        """Extract plain text from a document file.

        Supports .txt, .csv (direct read), .pdf (PyPDF2), .docx (python-docx).
        """
        if ext in (".txt", ".csv"):
            with open(filepath, "r", errors="replace") as f:
                return f.read()

        if ext == ".pdf":
            import PyPDF2
            text_parts = []
            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
            return "\n".join(text_parts)

        if ext == ".docx":
            import docx
            doc = docx.Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs if p.text)

        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts)

        return ""

    STOP_WORDS = frozenset({
        "the", "is", "a", "an", "what", "how", "do", "does", "are", "was",
        "were", "be", "to", "of", "and", "in", "for", "on", "with", "at",
        "by", "it", "i", "my", "me", "our", "this", "that", "can", "will",
    })

    def get_relevant_context(self, user_query, max_chars=12000):
        """Retrieve knowledge base context relevant to a user query.

        Uses keyword matching: splits the query into words, removes stop
        words, then scores each document by how many keywords appear in
        its extracted text. Returns the top-scoring documents formatted
        as labeled sections, capped at max_chars.
        """
        index = self._read_index()
        if not index:
            return ""

        # Build keyword set from the query
        keywords = set(user_query.lower().split()) - self.STOP_WORDS
        if not keywords:
            keywords = set(user_query.lower().split())

        # Score each document by keyword matches
        scored = []
        for entry in index:
            text_path = os.path.join(Config.EXTRACTED_TEXT_DIR, f"{entry['id']}.txt")
            if not os.path.exists(text_path):
                continue
            with open(text_path, "r") as f:
                text = f.read()
            text_lower = text.lower()
            score = sum(1 for kw in keywords if kw in text_lower)
            scored.append((entry, text, score))

        # Sort by score descending
        scored.sort(key=lambda item: item[2], reverse=True)

        # Build output up to max_chars
        result = ""
        for entry, text, score in scored:
            if score == 0:
                continue
            section = f"=== DOCUMENT: {entry['original_name']} ===\n{text}\n\n"
            if len(result) + len(section) > max_chars:
                break
            result += section

        return result

    def list_documents(self):
        """Return all documents in the knowledge base with metadata."""
        return self._read_index()

    def remove_document(self, doc_id):
        """Remove a document from the knowledge base. Returns True if removed."""
        index = self._read_index()
        entry = next((e for e in index if e["id"] == doc_id), None)
        if not entry:
            return False

        # Remove files
        doc_path = os.path.join(Config.KNOWLEDGE_BASE_DIR, entry["stored_name"])
        text_path = os.path.join(Config.EXTRACTED_TEXT_DIR, f"{doc_id}.txt")
        if os.path.exists(doc_path):
            os.remove(doc_path)
        if os.path.exists(text_path):
            os.remove(text_path)

        # Update index
        index = [e for e in index if e["id"] != doc_id]
        self._write_index(index)
        return True
